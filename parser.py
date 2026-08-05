"""
parser del archivo .xlsm/.xlsx de Seguimiento al PAC 2026 (INETER).
L
"""
import re
from datetime import datetime
from openpyxl import load_workbook

SHEET_NAME = "PAC"

# encabezados esperados -> nombre interno usado por el dashboard
COLUMN_MAP = {
    "SUB-UNIDAD": "subUnidad",
    "AREA SOLICITANTE": "areaSolicitante",
    "#Linea": "linea",
    "CBS": "cbs",
    "DESCRIPCION CONTRATACION": "descripcion",
    "MONTO ESTIMADO": "montoEstimado",
    "Estado": "estado",
    "EstadoDet": "estadoDet",
    "EstadoR": "estadoR",
    "Alerta": "alerta",
    "MODALIDAD": "modalidad",
    "FUENTE FINANCIA.": "fuenteFinancia",
    "F.PUBLICA": "fPublica",
    "Mes1": "mes",
    "F.RECEPCION": "fRecepcion",
    "F.EVALUA": "fEvalua",
    "F.ADJUDICA": "fAdjudica",
    "F.CONTRATO": "fContrato",
    "Fondos": "fondos",
    "OBSERVACION": "observacion",
}

MES_NORMALIZE = {
    "ene": "ene", "enero": "ene",
    "feb": "feb", "febrero": "feb",
    "mar": "mar", "marzo": "mar",
    "abr": "abr", "abril": "abr",
    "may": "may", "mayo": "may",
    "jun": "jun", "junio": "jun",
    "jul": "jul", "julio": "jul",
    "ago": "ago", "agosto": "ago",
    "sep": "sep", "sept": "sep", "septiembre": "sep", "set": "sep",
    "oct": "oct", "octubre": "oct",
    "nov": "nov", "noviembre": "nov",
    "dic": "dic", "diciembre": "dic",
}

# normalizaciones de texto pedidas explícitamente en Instrucciones.docx
TEXT_FIXES = {
    "LICITACIONSELECTIVA": "LICITACION SELECTIVA",
    "GobiernodeNicaragua": "Gobierno de Nicaragua",
}


def _clean_text(value):
    if value is None:
        return ""
    text = str(value).strip()
    for wrong, right in TEXT_FIXES.items():
        if text.replace(" ", "") == wrong.replace(" ", ""):
            return right
    # colapsa espacios repetidos que vienen de Excel
    text = re.sub(r"\s+", " ", text)
    return text


def _clean_money(value):
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value)
    text = re.sub(r"[^\d,.-]", "", text)
    text = text.replace(",", "")
    try:
        return float(text)
    except ValueError:
        return 0.0


def _clean_date(value):
    """Devuelve fecha ISO yyyy-mm-dd o '' si no hay fecha."""
    if value is None or value == "":
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    text = str(value).strip()
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return ""


def _clean_mes(value):
    if value is None:
        return ""
    text = str(value).strip().lower()
    return MES_NORMALIZE.get(text, text)


def parse_workbook(filepath):
    wb = load_workbook(filepath, data_only=True, read_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(
            f"No encontré la hoja '{SHEET_NAME}' en el archivo. "
            f"Hojas disponibles: {', '.join(wb.sheetnames)}"
        )
    ws = wb[SHEET_NAME]

    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    header = [str(h).strip() if h is not None else "" for h in header]

    idx_map = {}
    for i, col_name in enumerate(header):
        if col_name in COLUMN_MAP:
            idx_map[COLUMN_MAP[col_name]] = i

    missing = [c for c in COLUMN_MAP if COLUMN_MAP[c] not in idx_map]
    if missing:
        # No es fatal — seguimos con lo que haya, pero avisamos.
        print(f"Aviso: columnas no encontradas en el archivo: {missing}")

    records = []
    for row in rows:
        if row is None or all(v is None for v in row):
            continue

        def get(field):
            i = idx_map.get(field)
            return row[i] if i is not None and i < len(row) else None

        linea = get("linea")
        descripcion = _clean_text(get("descripcion"))
        if not descripcion and linea is None:
            continue  # fila vacía

        record = {
            "subUnidad": _clean_text(get("subUnidad")),
            "areaSolicitante": _clean_text(get("areaSolicitante")),
            "linea": linea if linea is not None else "",
            "cbs": _clean_text(get("cbs")),
            "descripcion": descripcion,
            "montoEstimado": _clean_money(get("montoEstimado")),
            "estado": _clean_text(get("estado")),
            "estadoDet": _clean_text(get("estadoDet")),
            "estadoR": _clean_text(get("estadoR")),
            "alerta": _clean_text(get("alerta")),
            "modalidad": _clean_text(get("modalidad")),
            "fuenteFinancia": _clean_text(get("fuenteFinancia")),
            "fPublica": _clean_date(get("fPublica")),
            "mes": _clean_mes(get("mes")),
            "fRecepcion": _clean_date(get("fRecepcion")),
            "fEvalua": _clean_date(get("fEvalua")),
            "fAdjudica": _clean_date(get("fAdjudica")),
            "fContrato": _clean_date(get("fContrato")),
            "fondos": _clean_text(get("fondos")),
            "observacion": _clean_text(get("observacion")),
        }
        records.append(record)

    wb.close()
    return records


if __name__ == "__main__":
    import sys
    import json
    path = sys.argv[1] if len(sys.argv) > 1 else "sample.xlsm"
    data = parse_workbook(path)
    print(f"{len(data)} registros parseados")
    print(json.dumps(data[:2], ensure_ascii=False, indent=2))
